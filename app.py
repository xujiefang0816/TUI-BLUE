from flask import Flask, render_template, redirect, url_for, flash, request, jsonify, make_response
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, SubmitField, SelectField, TextAreaField, FloatField, DateField
from wtforms.validators import DataRequired, Length, EqualTo
from functools import wraps
import datetime
from io import BytesIO
import json
from config import config
from models import db, User, FileType, Department, Unit, Status, File, OperationLog
from openpyxl import Workbook

# 创建应用实例
app = Flask(__name__)
app.config.from_object(config['default'])

# 初始化数据库
db.init_app(app)

# 初始化登录管理器
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = '请先登录后再访问该页面'

# 定义权限装饰器
def role_required(role):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            if current_user.role != role and current_user.role != 'admin':
                flash('您没有权限访问此页面')
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# 记录操作日志的函数
def log_operation(operation, details=''):
    if current_user.is_authenticated:
        log = OperationLog(
            user_id=current_user.id,
            operation=operation,
            details=details
        )
        db.session.add(log)
        db.session.commit()

# 登录加载用户
@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))



# 文件登记页面
@app.route('/file_register', methods=['GET', 'POST'])
@login_required
def file_register():
    file_types = FileType.query.all()
    departments = Department.query.all()
    units = Unit.query.all()
    
    if request.method == 'POST':
        try:
            date_str = request.form.get('date')
            date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
            
            file_type_id = request.form.get('file_type')
            department_id = request.form.get('department')
            applicant = request.form.get('applicant')
            
            # 处理付款相关文件的特殊情况
            file_type = FileType.query.get(file_type_id)
            content = request.form.get('content')
            
            # 检查是否是付款相关文件类型
            if file_type.name in ['付款申请单', '付款单+用印审批（仅限验收报告）']:
                payment_type = request.form.get('payment_type')
                payment_content = request.form.get('payment_content')
                payment_percentage = request.form.get('payment_percentage')
                payment_period = request.form.get('payment_period')
                payment_company = request.form.get('payment_company')
                
                # 合并摘要信息
                content = f"{payment_type} - {payment_content} (付款比例: {payment_percentage}%, 期间: {payment_period}, 单位: {payment_company})"
            
            unit_id = request.form.get('unit')
            quantity = float(request.form.get('quantity', 0))
            amount = float(request.form.get('amount', 0))
            
            # 创建新文件记录
            new_file = File(
                date=date,
                file_type_id=file_type_id,
                department_id=department_id,
                applicant=applicant,
                content=content,
                unit_id=unit_id,
                quantity=quantity,
                amount=amount,
                created_by=current_user.id
            )
            
            db.session.add(new_file)
            db.session.commit()
            
            log_operation('登记文件', f'文件类型: {file_type.name}, 申请人: {applicant}')
            flash('文件登记成功！')
            return redirect(url_for('file_information'))
        except Exception as e:
            db.session.rollback()
            flash(f'文件登记失败: {str(e)}')
    
    return render_template('file_register.html', 
                          file_types=file_types,
                          departments=departments,
                          units=units)







# 删除文件
@app.route('/delete_file/<int:file_id>')




# 批量更新状态
@app.route('/batch_update_status', methods=['POST'])
@role_required('manager')
def batch_update_status():
    try:
        file_ids = request.form.getlist('file_ids[]')
        status_id = request.form.get('status_id')
        
        if not status_id:
            flash('请选择状态！')
            return redirect(url_for('file_processing'))
        
        status = Status.query.get(status_id)
        
        for file_id in file_ids:
            file = File.query.get(int(file_id))
            file.status_id = status_id
            file.submission_time = datetime.datetime.now()
            
            # 如果状态是"完毕"，设置结束日期
            if status.name == '完毕':
                file.end_date = datetime.date.today()
        
        db.session.commit()
        
        log_operation('批量更新状态', f'更新文件数量: {len(file_ids)}, 状态: {status.name}')
        flash(f'成功更新{len(file_ids)}个文件的状态！')
    except Exception as e:
        db.session.rollback()
        flash(f'批量更新失败: {str(e)}')
    
    return redirect(url_for('file_processing'))

# 导出Excel
@app.route('/export_excel')
@role_required('manager')
def export_excel():
    files = File.query.all()
    
    # 创建Excel工作簿
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '文件信息'
    
    # 设置表头
    headers = ['日期', '文件类型', '文件编号', '申请部门', '申请人', '文件内容', 
               '计量单位', '数量', '金额', '结束日期', '送签状态', '退回原因', '备注', '送签时间']
    for col_idx, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_idx, value=header)
    
    # 填充数据
    for row_idx, file in enumerate(files, 2):
        sheet.cell(row=row_idx, column=1, value=file.date.strftime('%Y-%m-%d') if file.date else '')
        sheet.cell(row=row_idx, column=2, value=file.file_type.name if file.file_type else '')
        sheet.cell(row=row_idx, column=3, value=file.file_number or '')
        sheet.cell(row=row_idx, column=4, value=file.department.name if file.department else '')
        sheet.cell(row=row_idx, column=5, value=file.applicant)
        sheet.cell(row=row_idx, column=6, value=file.content or '')
        sheet.cell(row=row_idx, column=7, value=file.unit.name if file.unit else '')
        sheet.cell(row=row_idx, column=8, value=file.quantity if file.quantity != 0 else '/')
        sheet.cell(row=row_idx, column=9, value=file.amount if file.amount != 0 else '/')
        sheet.cell(row=row_idx, column=10, value=file.end_date.strftime('%Y-%m-%d') if file.end_date else '')
        sheet.cell(row=row_idx, column=11, value=file.status.name if file.status else '')
        sheet.cell(row=row_idx, column=12, value=file.return_reason or '')
        sheet.cell(row=row_idx, column=13, value=file.remark or '')
        sheet.cell(row=row_idx, column=14, value=file.submission_time.strftime('%Y-%m-%d %H:%M:%S') if file.submission_time else '')
    
    # 创建临时文件
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    
    # 生成响应
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=文件信息_' + datetime.datetime.now().strftime('%Y%m%d%H%M%S') + '.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    log_operation('导出Excel', '导出所有文件信息')
    
    return response



# 添加文件类型
@app.route('/add_file_type', methods=['POST'])
@role_required('admin')
def add_file_type():
    name = request.form.get('name')
    
    if FileType.query.filter_by(name=name).first():
        flash('该文件类型已存在！')
    else:
        try:
            new_type = FileType(name=name)
            db.session.add(new_type)
            db.session.commit()
            
            log_operation('添加文件类型', name)
            flash('文件类型添加成功！')
        except Exception as e:
            db.session.rollback()
            flash(f'添加失败: {str(e)}')
    
    return redirect(url_for('system_settings'))

# 删除文件类型
@app.route('/delete_file_type/<int:type_id>')
@role_required('admin')
def delete_file_type(type_id):
    file_type = FileType.query.get_or_404(type_id)
    
    # 检查是否有文件使用该类型
    if File.query.filter_by(file_type_id=type_id).first():
        flash('该文件类型有文件使用，无法删除！')
    else:
        try:
            db.session.delete(file_type)
            db.session.commit()
            
            log_operation('删除文件类型', file_type.name)
            flash('文件类型删除成功！')
        except Exception as e:
            db.session.rollback()
            flash(f'删除失败: {str(e)}')
    
    return redirect(url_for('system_settings'))

# 添加部门
@app.route('/add_department', methods=['POST'])
@role_required('admin')
def add_department():
    name = request.form.get('name')
    
    if Department.query.filter_by(name=name).first():
        flash('该部门已存在！')
    else:
        try:
            new_dept = Department(name=name)
            db.session.add(new_dept)
            db.session.commit()
            
            log_operation('添加部门', name)
            flash('部门添加成功！')
        except Exception as e:
            db.session.rollback()
            flash(f'添加失败: {str(e)}')
    
    return redirect(url_for('system_settings'))

# 删除部门
@app.route('/delete_department/<int:dept_id>')
@role_required('admin')
def delete_department(dept_id):
    department = Department.query.get_or_404(dept_id)
    
    # 检查是否有文件使用该部门
    if File.query.filter_by(department_id=dept_id).first():
        flash('该部门有文件使用，无法删除！')
    else:
        try:
            db.session.delete(department)
            db.session.commit()
            
            log_operation('删除部门', department.name)
            flash('部门删除成功！')
        except Exception as e:
            db.session.rollback()
            flash(f'删除失败: {str(e)}')
    
    return redirect(url_for('system_settings'))

# 添加计量单位
@app.route('/add_unit', methods=['POST'])
@role_required('admin')
def add_unit():
    name = request.form.get('name')
    
    if Unit.query.filter_by(name=name).first():
        flash('该计量单位已存在！')
    else:
        try:
            new_unit = Unit(name=name)
            db.session.add(new_unit)
            db.session.commit()
            
            log_operation('添加计量单位', name)
            flash('计量单位添加成功！')
        except Exception as e:
            db.session.rollback()
            flash(f'添加失败: {str(e)}')
    
    return redirect(url_for('system_settings'))

# 删除计量单位
@app.route('/delete_unit/<int:unit_id>')
@role_required('admin')
def delete_unit(unit_id):
    unit = Unit.query.get_or_404(unit_id)
    
    # 检查是否有文件使用该计量单位
    if File.query.filter_by(unit_id=unit_id).first():
        flash('该计量单位有文件使用，无法删除！')
    else:
        try:
            db.session.delete(unit)
            db.session.commit()
            
            log_operation('删除计量单位', unit.name)
            flash('计量单位删除成功！')
        except Exception as e:
            db.session.rollback()
            flash(f'删除失败: {str(e)}')
    
    return redirect(url_for('system_settings'))

# 添加送签状态
@app.route('/add_status', methods=['POST'])
@role_required('admin')
def add_status():
    name = request.form.get('name')
    
    if Status.query.filter_by(name=name).first():
        flash('该送签状态已存在！')
    else:
        try:
            new_status = Status(name=name)
            db.session.add(new_status)
            db.session.commit()
            
            log_operation('添加送签状态', name)
            flash('送签状态添加成功！')
        except Exception as e:
            db.session.rollback()
            flash(f'添加失败: {str(e)}')
    
    return redirect(url_for('system_settings'))

# 删除送签状态
@app.route('/delete_status/<int:status_id>')
@role_required('admin')
def delete_status(status_id):
    status = Status.query.get_or_404(status_id)
    
    # 检查是否有文件使用该状态
    if File.query.filter_by(status_id=status_id).first():
        flash('该送签状态有文件使用，无法删除！')
    else:
        try:
            db.session.delete(status)
            db.session.commit()
            
            log_operation('删除送签状态', status.name)
            flash('送签状态删除成功！')
        except Exception as e:
            db.session.rollback()
            flash(f'删除失败: {str(e)}')
    
    return redirect(url_for('system_settings'))

# 添加用户
@app.route('/add_user', methods=['POST'])
@role_required('admin')
def add_user():
    username = request.form.get('username')
    password = request.form.get('password')
    role = request.form.get('role')
    
    if User.query.filter_by(username=username).first():
        flash('该用户名已存在！')
    else:
        try:
            new_user = User(username=username, role=role)
            new_user.set_password(password)
            db.session.add(new_user)
            db.session.commit()
            
            log_operation('添加用户', f'用户名: {username}, 角色: {role}')
            flash('用户添加成功！')
        except Exception as e:
            db.session.rollback()
            flash(f'添加失败: {str(e)}')
    
    return redirect(url_for('system_settings'))

# 删除用户
@app.route('/delete_user/<int:user_id>')
@role_required('admin')
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    
    # 不能删除当前登录用户
    if user.id == current_user.id:
        flash('不能删除当前登录的用户！')
    else:
        try:
            db.session.delete(user)
            db.session.commit()
            
            log_operation('删除用户', f'用户名: {user.username}, 角色: {user.role}')
            flash('用户删除成功！')
        except Exception as e:
            db.session.rollback()
            flash(f'删除失败: {str(e)}')
    
    return redirect(url_for('system_settings'))

# 操作日志页面
@app.route('/operation_logs')
@role_required('admin')
def operation_logs():
    logs = OperationLog.query.order_by(OperationLog.timestamp.desc()).all()
    return render_template('operation_logs.html', logs=logs)



# 主函数
if __name__ == '__main__':
    # 在应用上下文中创建数据库表
    with app.app_context():
        db.create_all()
    app.run(debug=True)

# 表单类
class LoginForm(FlaskForm):
    username = StringField('用户名', validators=[DataRequired(), Length(min=2, max=20)])
    password = PasswordField('密码', validators=[DataRequired()])
    submit = SubmitField('登录')

class RegisterForm(FlaskForm):
    username = StringField('用户名', validators=[DataRequired(), Length(min=2, max=20)])
    password = PasswordField('密码', validators=[DataRequired()])
    confirm_password = PasswordField('确认密码', validators=[DataRequired(), EqualTo('password')])
    role = SelectField('角色', choices=[('user', '普通账号'), ('manager', '普通管理员'), ('admin', '总管理员')], default='user')
    submit = SubmitField('注册')

    def validate_username(self, username):
        user = User.query.filter_by(username=username.data).first()
        if user:
            raise ValidationError('该用户名已被使用，请选择其他用户名。')

class ChangePasswordForm(FlaskForm):
    current_password = PasswordField('当前密码', validators=[DataRequired()])
    new_password = PasswordField('新密码', validators=[DataRequired()])
    confirm_new_password = PasswordField('确认新密码', validators=[DataRequired(), EqualTo('new_password')])
    submit = SubmitField('修改密码')

class FileForm(FlaskForm):
    date = DateField('日期', validators=[DataRequired()], default=datetime.date.today)
    file_type = SelectField('文件类型', coerce=int, validators=[DataRequired()])
    department = SelectField('部门', coerce=int, validators=[DataRequired()])
    applicant = StringField('申请人', validators=[DataRequired(), Length(max=100)])
    content = TextAreaField('文件内容', validators=[Length(max=1000)])
    unit = SelectField('计量单位', coerce=int, validators=[DataRequired()])
    quantity = FloatField('数量')
    amount = FloatField('金额')
    submit = SubmitField('提交文件')
    reset = SubmitField('重置')

class SearchForm(FlaskForm):
    search = StringField('搜索', render_kw={'placeholder': '搜索文件信息...'})
    submit = SubmitField('搜索')

class StatusForm(FlaskForm):
    status = SelectField('送签状态', coerce=int, validators=[DataRequired()])
    submit = SubmitField('确认')

# 路由
@app.route('/')
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user and user.password == encrypt_password(form.password.data):
            login_user(user)
            # 记录登录日志
            log = OperationLog(user_id=user.id, operation='登录', details=f'用户 {user.username} 登录系统')
            db.session.add(log)
            db.session.commit()
            return redirect(url_for('index'))
        else:
            flash('登录失败，请检查用户名和密码')
    return render_template('login.html', form=form)

@app.route('/logout')
@login_required
def logout():
    # 记录登出日志
    log = OperationLog(user_id=current_user.id, operation='登出', details=f'用户 {current_user.username} 登出系统')
    db.session.add(log)
    db.session.commit()
    logout_user()
    return redirect(url_for('login'))

@app.route('/index')
@login_required
def index():
    return render_template('index.html')



@app.route('/file_information', methods=['GET', 'POST'])
@login_required
def file_information():
    search_form = SearchForm()
    files = File.query.all()
    
    if search_form.validate_on_submit():
        search_term = search_form.search.data.lower()
        files = File.query.join(FileType).join(Department).join(Unit).join(User).filter(
            (File.date.cast(db.String).contains(search_term)) |
            (FileType.name.contains(search_term)) |
            (Department.name.contains(search_term)) |
            (File.applicant.contains(search_term)) |
            (File.content.contains(search_term)) |
            (Unit.name.contains(search_term)) |
            (File.quantity.cast(db.String).contains(search_term)) |
            (File.amount.cast(db.String).contains(search_term)) |
            (User.username.contains(search_term))
        ).all()
    
    # 记录查看日志
    if request.method == 'GET' and 'no_log' not in request.args:
        log = OperationLog(user_id=current_user.id, operation='查看文件信息', details=f'用户 {current_user.username} 查看了文件信息列表')
        db.session.add(log)
        db.session.commit()
    
    return render_template('file_information.html', files=files, search_form=search_form)

@app.route('/file_processing', methods=['GET', 'POST'])
@login_required
@role_required('manager')
def file_processing():
    search_form = SearchForm()
    files = File.query.all()
    
    if search_form.validate_on_submit():
        search_term = search_form.search.data.lower()
        files = File.query.join(FileType).join(Department).join(Unit).join(User).filter(
            (File.date.cast(db.String).contains(search_term)) |
            (FileType.name.contains(search_term)) |
            (Department.name.contains(search_term)) |
            (File.applicant.contains(search_term)) |
            (File.content.contains(search_term)) |
            (Unit.name.contains(search_term)) |
            (File.quantity.cast(db.String).contains(search_term)) |
            (File.amount.cast(db.String).contains(search_term)) |
            (User.username.contains(search_term))
        ).all()
    
    # 记录操作日志
    if request.method == 'GET' and 'no_log' not in request.args:
        log = OperationLog(user_id=current_user.id, operation='进入文件处理', details=f'用户 {current_user.username} 进入了文件处理页面')
        db.session.add(log)
        db.session.commit()
    
    return render_template('file_processing.html', files=files, search_form=search_form)

@app.route('/edit_file/<int:file_id>', methods=['GET', 'POST'])
@login_required
@role_required('manager')
def edit_file(file_id):
    file = File.query.get_or_404(file_id)
    form = FileForm(obj=file)
    form.file_type.choices = [(ft.id, ft.name) for ft in FileType.query.all()]
    form.department.choices = [(d.id, d.name) for d in Department.query.all()]
    form.unit.choices = [(u.id, u.name) for u in Unit.query.all()]
    
    if form.validate_on_submit():
        form.populate_obj(file)
        
        # 记录操作日志
        log = OperationLog(user_id=current_user.id, operation='编辑文件', details=f'用户 {current_user.username} 编辑了文件 ID: {file_id}')
        db.session.add(log)
        
        db.session.commit()
        flash('文件已更新')
        return redirect(url_for('file_processing'))
    
    return render_template('edit_file.html', form=form, file=file)

@app.route('/delete_file/<int:file_id>')
@login_required
@role_required('manager')
def delete_file(file_id):
    file = File.query.get_or_404(file_id)
    
    # 记录操作日志
    log = OperationLog(user_id=current_user.id, operation='删除文件', details=f'用户 {current_user.username} 删除了文件 ID: {file_id}')
    db.session.add(log)
    
    db.session.delete(file)
    db.session.commit()
    flash('文件已删除')
    return redirect(url_for('file_processing'))

@app.route('/batch_delete_files', methods=['POST'])
@login_required
@role_required('manager')
def batch_delete_files():
    file_ids = request.form.getlist('file_ids[]')
    for file_id in file_ids:
        file = File.query.get_or_404(int(file_id))
        db.session.delete(file)
    
    # 记录操作日志
    log = OperationLog(user_id=current_user.id, operation='批量删除文件', details=f'用户 {current_user.username} 批量删除了 {len(file_ids)} 个文件')
    db.session.add(log)
    
    db.session.commit()
    flash(f'已成功删除 {len(file_ids)} 个文件')
    return redirect(url_for('file_processing'))

@app.route('/update_status/<int:file_id>', methods=['POST'])
@login_required
@role_required('manager')
def update_status(file_id):
    file = File.query.get_or_404(file_id)
    status_id = request.form.get('status_id')
    
    if status_id:
        file.status_id = int(status_id)
        if Status.query.get(int(status_id)).name == '完毕':
            file.end_date = datetime.date.today()
        
        # 记录操作日志
        log = OperationLog(user_id=current_user.id, operation='更新文件状态', details=f'用户 {current_user.username} 更新了文件 ID: {file_id} 的状态')
        db.session.add(log)
        
        db.session.commit()
    
    return redirect(url_for('file_processing'))





@app.route('/system_settings')
@login_required
@role_required('admin')
def system_settings():
    file_types = FileType.query.all()
    departments = Department.query.all()
    units = Unit.query.all()
    statuses = Status.query.all()
    users = User.query.all()
    
    # 记录操作日志
    log = OperationLog(user_id=current_user.id, operation='进入系统设置', details=f'用户 {current_user.username} 进入了系统设置页面')
    db.session.add(log)
    db.session.commit()
    
    return render_template('system_settings.html', 
                          file_types=file_types, 
                          departments=departments, 
                          units=units, 
                          statuses=statuses, 
                          users=users)





















@app.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    form = ChangePasswordForm()
    
    if form.validate_on_submit():
        # 验证当前密码
        if current_user.password != encrypt_password(form.current_password.data):
            flash('当前密码不正确')
            return redirect(url_for('change_password'))
        
        # 更新密码
        current_user.password = encrypt_password(form.new_password.data)
        
        # 记录操作日志
        log = OperationLog(user_id=current_user.id, operation='修改密码', details=f'用户 {current_user.username} 修改了自己的密码')
        db.session.add(log)
        
        db.session.commit()
        flash('密码已更新')
        return redirect(url_for('index'))
    
    return render_template('change_password.html', form=form)



@app.route('/clear_logs')
@login_required
@role_required('admin')
def clear_logs():
    # 记录清理日志
    log = OperationLog(user_id=current_user.id, operation='清空日志', details=f'用户 {current_user.username} 清空了所有操作日志')
    db.session.add(log)
    db.session.commit()
    
    # 删除所有日志（除了刚刚添加的这条）
    OperationLog.query.filter(OperationLog.id != log.id).delete()
    db.session.commit()
    
    flash('操作日志已清空')
    return redirect(url_for('operation_logs'))

# 初始化数据库和默认数据
def init_db():
    with app.app_context():
        db.create_all()
        
        # 添加默认文件类型
        default_file_types = [
            '采购计划审批表', '合同（协议）签订审批表', '付款申请单', '用印审批表',
            '付款单+用印审批（仅限验收报告）', '工作联系单', '固定资产验收单',
            '会议议题', '借印审批表', '请假申请表', '差旅申请表', '其他'
        ]
        
        for type_name in default_file_types:
            if not FileType.query.filter_by(name=type_name).first():
                db.session.add(FileType(name=type_name))
        
        # 添加默认部门
        default_departments = [
            '前厅FO', '客房HSKP', '西餐厅', '中餐厅', '大堂吧', '宴会厅',
            '迷你吧', '餐饮办公室', '管事部', '饼房', '财务FIN', '行政EO',
            '人事HR', '员工餐厅', '销售S&M', '工程ENG'
        ]
        
        for dept_name in default_departments:
            if not Department.query.filter_by(name=dept_name).first():
                db.session.add(Department(name=dept_name))
        
        # 添加默认计量单位
        default_units = [
            '/', '批', '个（支）', '件', '套', '份', '只', '台', '桶', '次',
            '块', '人', '盒', '瓶', '双', '张', '月', '年', '克（g）',
            '千克（kg）', '箱', '米', '平方米', '包', '袋', '家', 'PCS',
            'PAC', '佣金（%）', '其他'
        ]
        
        for unit_name in default_units:
            if not Unit.query.filter_by(name=unit_name).first():
                db.session.add(Unit(name=unit_name))
        
        # 添加默认送签状态
        default_statuses = [
            '完毕', '总秘（酒店总经理）', '待送集团', '业主代表',
            '陆总及彭总（盖章处）', '集团审核', '集团经理待签', '退回',
            '未盖章', '重签', '作废', '资产管理部', '招采办',
            '酒店内部走签', '急单', '已签未付'
        ]
        
        for status_name in default_statuses:
            if not Status.query.filter_by(name=status_name).first():
                db.session.add(Status(name=status_name))
        
        # 添加默认用户
        # 总管理员账号 TYL2025，密码 941314aA
        if not User.query.filter_by(username='TYL2025').first():
            admin_user = User(username='TYL2025', password=encrypt_password('941314aA'), role='admin')
            db.session.add(admin_user)
        
        # 普通管理员账号 8888，密码 8888
        if not User.query.filter_by(username='8888').first():
            manager_user = User(username='8888', password=encrypt_password('8888'), role='manager')
            db.session.add(manager_user)
        
        # 普通账号 1001，密码 1001
        if not User.query.filter_by(username='1001').first():
            normal_user = User(username='1001', password=encrypt_password('1001'), role='user')
            db.session.add(normal_user)
        
        db.session.commit()

# 运行应用
if __name__ == '__main__':
    init_db()
    app.run(debug=True)